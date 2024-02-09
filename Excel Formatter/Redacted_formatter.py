import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy
import subprocess

def create_gui():
    root = tk.Tk() # Create the root window
    root.title("TASK Formatter") # Set the title bar text
    
    selected_file = None  # Variable to store the selected file path
    
    # File selection button
    def select_file():
        nonlocal selected_file
        filepath = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if filepath.endswith('.csv'):
            filename = os.path.basename(filepath)
            file_label.config(text=filename)
            selected_file = filepath
            format_button.config(state=tk.NORMAL)  # Enable the format button

        # Hide the formatted label
        formatted_label.grid_remove()

    def format_sheet(sheet):
        # Wrap all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Set column width to 75
        for column in sheet.columns:
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = 20

        # Set row height to 50
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 45

        # Set the header row to bold
        for cell in sheet[1]:
            new_font = copy(cell.font)
            new_font.bold = True
            cell.font = new_font

        # Set a filter on the first row
        sheet.auto_filter.ref = sheet.dimensions

    def format_and_save_file(file_path, columns_to_keep, output_filename):
        # Remove columns not in the list
        df = pd.read_csv(file_path)
        df = df[columns_to_keep]
        df.to_excel(output_filename, index=False)

        # Load the workbook and get the first worksheet
        wb = load_workbook(filename=output_filename)
        sheet = wb.active   

        # Format the sheet
        format_sheet(sheet)

        # Save the changes, overwriting the file if it already exists
        wb.save(filename=output_filename)

        # Show the formatted label
        formatted_label.grid()

        # [OPTIONAL] Open the formatted file location in file explorer
        #subprocess.Popen(r'explorer /select,"{}"'.format(output_filename.replace('/', '\\')))

    # Format file button
    def format_file():

        # If a file has been selected
        if selected_file:
            # Save output file name
            output_filename = os.path.splitext(selected_file)[0] + "_Formatted.xlsx"

            if selected_choice.get() == "Selection 01":
                # Define the list of columns to keep
                columns_to_keep_select1 = [
                    # List of columns to keep
                ]

                format_and_save_file(selected_file, columns_to_keep_select1, output_filename)
                pass
            elif selected_choice.get() == "Selection 02":
                # Define the list of columns to keep
                columns_to_keep_select2 = [
                    # List of columns to keep
                ]

                format_and_save_file(selected_file, columns_to_keep_select2, output_filename)
                pass
            elif selected_choice.get() == "Selection 03":
                # Define the list of columns to keep
                columns_to_keep_select3 = [
                    # List of columns to keep
                ]

                format_and_save_file(selected_file, columns_to_keep_select3, output_filename)
                pass 
    
    # Button to select a file
    select_button = ttk.Button(root, text="Select File", command=select_file)
    select_button.grid(row=1, column=0, padx=10, pady=0)

    # Label to display the file name chosen
    file_label = tk.Label(root, text="")
    file_label.grid(row=0, column=0, padx=10, pady=30)  # Update row and column parameters

    # Combobox to select a choice
    choices = ["Selection 01", "Selection 02", "Selection 03"]
    selected_choice = tk.StringVar()
    choice_combobox = ttk.Combobox(root, textvariable=selected_choice, values=choices, state="readonly")
    choice_combobox.grid(row=2, column=0, padx=10, pady=10)

    # Button to format the selected file
    format_button = ttk.Button(root, text="Format File", command=format_file, state=tk.DISABLED)  # Disable the format button initially
    format_button.grid(row=3, column=0, padx=10, pady=0)

    # Label to display when file has been formatted
    formatted_label = tk.Label(root, text="Formatted!", fg="green")
    formatted_label.grid(row=4, column=0, padx=10, pady=10)
    formatted_label.grid_remove()  # Hide the label initially

    # Function to enable/disable the format button based on file selection and choice selection
    def update_format_button_state():
        if selected_file and selected_choice.get():
            format_button.config(state=tk.NORMAL)
        else:
            format_button.config(state=tk.DISABLED)

    # Update the format button state whenever the file selection or choice selection changes
    file_label.bind("<Configure>", lambda event: update_format_button_state())
    choice_combobox.bind("<<ComboboxSelected>>", lambda event: update_format_button_state())
    
    # Configure grid weights to make the widgets expand and fill the available space
    root.grid_rowconfigure(0, weight=0)
    root.grid_columnconfigure(0, weight=1)
    
    # Set the window size
    root.geometry("400x250") 
    
    # Run the GUI
    root.mainloop()

create_gui()